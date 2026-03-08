import streamlit as st
import pandas as pd
import io
from deep_translator import GoogleTranslator
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ==========================================
# 1. CONFIGURAÇÃO E MEMÓRIA
# ==========================================
st.set_page_config(page_title="SpeedQuote Global", page_icon="⚡", layout="wide")

if 'logged_in' not in st.session_state: st.session_state['logged_in'] = False
if 'current_page' not in st.session_state: st.session_state['current_page'] = 'Dashboard'

def change_page(page_name): st.session_state['current_page'] = page_name

# ==========================================
# 2. INJEÇÃO DE CSS
# ==========================================
st.markdown("""
    <style>
    .stApp { background-color: #050505; color: #ffffff; }
    .neon-orb-purple { position: absolute; width: 500px; height: 500px; background: #8A2BE2; border-radius: 50%; filter: blur(180px); top: 0%; right: 10%; z-index: 0; opacity: 0.4; animation: pulse 8s infinite alternate; }
    .neon-orb-green { position: absolute; width: 400px; height: 400px; background: #00ff41; border-radius: 50%; filter: blur(150px); bottom: -10%; left: 10%; z-index: 0; opacity: 0.2; }
    @keyframes pulse { 0% { transform: scale(1); opacity: 0.3; } 100% { transform: scale(1.1); opacity: 0.5; } }
    .sales-title { font-size: 3.8rem; font-weight: 900; background: linear-gradient(45deg, #b026ff, #00ff41); -webkit-background-clip: text; -webkit-text-fill-color: transparent; margin-bottom: 0px; line-height: 1.1; z-index: 1; position: relative; }
    .sales-subtitle { font-size: 1.4rem; color: #a0a0a0; margin-top: 10px; margin-bottom: 40px; z-index: 1; position: relative; }
    .bullet-point { font-size: 1.1rem; margin-bottom: 20px; color: #e0e0e0; z-index: 1; position: relative; line-height: 1.5; }
    .glass-box { background: rgba(20, 20, 20, 0.4); backdrop-filter: blur(16px); -webkit-backdrop-filter: blur(16px); border: 1px solid rgba(255, 255, 255, 0.08); border-radius: 20px; padding: 40px; box-shadow: 0 8px 32px 0 rgba(0, 0, 0, 0.6); z-index: 1; position: relative; }
    .btn-main>button { width: 100% !important; background-color: #00ff41 !important; color: #050505 !important; font-weight: 900 !important; font-size: 1.1rem !important; border: none !important; border-radius: 10px !important; padding: 12px !important; transition: all 0.3s ease !important; }
    .btn-main>button:hover { background-color: #00cc33 !important; box-shadow: 0 0 20px rgba(0, 255, 65, 0.6) !important; transform: scale(1.02); }
    .nav-container { background: linear-gradient(90deg, #11071F, #050505, #0A140D); padding: 10px; border-bottom: 1px solid rgba(138, 43, 226, 0.3); border-radius: 10px; margin-bottom: 30px; box-shadow: 0 4px 15px rgba(0,0,0,0.5); z-index: 2; position: relative; }
    .nav-btn>button { background: transparent !important; color: #a0a0a0 !important; border: none !important; font-weight: bold !important; font-size: 1.1rem !important; transition: 0.3s !important; }
    .nav-btn>button:hover { color: #00ff41 !important; transform: translateY(-2px); }
    .dash-card { background: rgba(20, 20, 20, 0.6); border: 1px solid rgba(255, 255, 255, 0.05); border-radius: 15px; padding: 25px; height: 100%; transition: transform 0.3s; position: relative; z-index: 1; }
    .dash-card:hover { transform: translateY(-5px); border-color: rgba(138, 43, 226, 0.4); box-shadow: 0 10px 20px rgba(0,0,0,0.5); }
    .dash-title { font-size: 1.2rem; font-weight: bold; color: #ffffff; margin-bottom: 10px; }
    .dash-value { font-size: 2.8rem; font-weight: 900; color: #8A2BE2; margin-bottom: 5px; }
    .tool-header { color: #00ff41; border-bottom: 2px solid #8A2BE2; padding-bottom: 10px; margin-bottom: 20px; }
    div[data-testid="InputInstructions"] > span:nth-child(1) { visibility: hidden; }
    </style>
    <div class="neon-orb-purple"></div><div class="neon-orb-green"></div>
""", unsafe_allow_html=True)

# ==========================================
# 3. MÁQUINA DE ESTADOS E ROTEAMENTO
# ==========================================
if not st.session_state['logged_in']:
    spacer_left, col_sales, col_login, spacer_right = st.columns([0.5, 2.5, 1.5, 0.5])
    with col_sales:
        st.write("<br><br><br>", unsafe_allow_html=True)
        st.markdown('<p class="sales-title">SPEEDQUOTE GLOBAL</p>', unsafe_allow_html=True)
        st.markdown('<p class="sales-subtitle">Não faça planilhas. Faça negócios.</p>', unsafe_allow_html=True)
        st.markdown('<p class="bullet-point">⚡ <b>Mapeamento Automático:</b> Traduza e formate dados crus em cotações padronizadas em segundos.</p>', unsafe_allow_html=True)
        st.markdown('<p class="bullet-point">🔗 <b>Smart Link:</b> Esqueça arquivos no WeChat. Envie links interativos para a China e receba os preços direto no painel.</p>', unsafe_allow_html=True)
        st.markdown('<p class="bullet-point">📦 <b>Hub Logístico:</b> Calcule o espaço no container (CBM) e evite prejuízos de <i>Inland Freight</i>. <span style="color:#FFD700;">🔒</span></p>', unsafe_allow_html=True)
    with col_login:
        st.write("<br><br>", unsafe_allow_html=True)
        st.markdown("<div class='glass-box'>", unsafe_allow_html=True)
        st.markdown("<h2 style='text-align: center; margin-bottom: 15px;'>Acesse o Portal</h2>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; color: #a0a0a0; font-size: 0.9em; margin-bottom: 25px;'><b>Login</b> &nbsp;|&nbsp; <a href='#' style='color: #8A2BE2; text-decoration: none;'>Sign Up</a></p>", unsafe_allow_html=True)
        email = st.text_input("E-mail corporativo", placeholder="voce@suaempresa.com")
        senha = st.text_input("Senha", type="password", placeholder="••••••••")
        st.markdown("<div class='btn-main'>", unsafe_allow_html=True)
        if st.button("INICIAR SESSÃO"):
            if email and senha:
                st.session_state['logged_in'] = True
                st.rerun()
            else: st.error("Preencha o e-mail e a senha.")
        st.markdown("</div></div>", unsafe_allow_html=True)
else:
    st.markdown('<div class="nav-container">', unsafe_allow_html=True)
    n1, n2, n3, n4, n5 = st.columns([1, 1.5, 1.5, 1.5, 1.5])
    with n1: st.markdown("<span style='font-size: 1.5rem; font-weight: 900; background: linear-gradient(45deg, #b026ff, #00ff41); -webkit-background-clip: text; -webkit-text-fill-color: transparent;'>⚡ SQG</span>", unsafe_allow_html=True)
    st.markdown("<div class='nav-btn'>", unsafe_allow_html=True)
    with n2: 
        if st.button("🏠 Dashboard"): change_page('Dashboard')
    with n3: 
        if st.button("⚡ Auto-Format"): change_page('Auto-Format')
    with n4: st.button("🔗 Smart Link 🔒", disabled=True)
    with n5: st.button("📦 Logística 🔒", disabled=True)
    st.markdown("</div></div>", unsafe_allow_html=True)
    
    if st.session_state['current_page'] == 'Dashboard':
        col_texto, col_botao = st.columns([8, 1])
        with col_texto: st.markdown("<h2 style='margin-bottom: 20px;'>Bem-vindo de volta, Importador.</h2>", unsafe_allow_html=True)
        with col_botao:
            if st.button("Sair"):
                st.session_state['logged_in'] = False
                st.rerun()
        c1, c2, c3 = st.columns(3)
        with c1: st.markdown("""<div class="dash-card"><div class="dash-title">📊 Uso da Conta (Plano Free)</div><div class="dash-value" style="color: #00ff41;">1 <span style="font-size: 1.2rem; color: #a0a0a0;">/ 5</span></div><p style="color: #a0a0a0; font-size: 0.9em;">Cotações gratuitas utilizadas neste mês.</p><div style="width: 100%; background-color: #222; height: 8px; border-radius: 5px; margin-top: 20px;"><div style="width: 20%; background-color: #00ff41; height: 100%; border-radius: 5px; box-shadow: 0 0 10px #00ff41;"></div></div></div>""", unsafe_allow_html=True)
        with c2:
            st.markdown("""<div class="dash-card" style="border-color: rgba(0, 255, 65, 0.4); background: rgba(0, 255, 65, 0.05);"><div class="dash-title">🚀 Ação Rápida</div><p style="color: #e0e0e0; font-size: 0.95em; margin-bottom: 20px;">Inicie uma formatação de planilha agora.</p></div>""", unsafe_allow_html=True)
            st.markdown("<div class='btn-main' style='margin-top: -60px; padding: 0 25px; position: relative; z-index: 5;'>", unsafe_allow_html=True)
            if st.button("⚡ NOVA COTAÇÃO AUTOMÁTICA"):
                change_page('Auto-Format')
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)
        with c3: st.markdown("""<div class="dash-card" style="padding: 0; overflow: hidden; border-color: rgba(255, 215, 0, 0.4);"><div style="filter: blur(5px); opacity: 0.2; height: 100%; position: absolute; width: 100%; background-image: url('https://upload.wikimedia.org/wikipedia/commons/thumb/e/e0/China_edcp_location_map.svg/500px-China_edcp_location_map.svg.png'); background-size: cover; background-position: center;"></div><div style="position: relative; z-index: 2; top: 50%; transform: translateY(30%); text-align: center; width: 100%;"><h3 style="color: #FFD700; margin-bottom: 5px;">🔒 Módulo Premium</h3><p style="font-size: 0.9em; color: #fff;">Calcule o CBM e Frete Inland.</p><button style="background: #FFD700; color: #000; border: none; padding: 10px 20px; border-radius: 5px; font-weight: 900; cursor: pointer; margin-top: 10px;">DESBLOQUEAR PLANO</button></div></div>""", unsafe_allow_html=True)

    elif st.session_state['current_page'] == 'Auto-Format':
        st.markdown("<h2 class='tool-header'>⚡ Auto-Format (Mapeamento Mágico)</h2>", unsafe_allow_html=True)
        
        st.markdown("<div class='glass-box'>", unsafe_allow_html=True)
        mapa_idiomas = {"🇺🇸 English": "en", "🇨🇳 中文 (Mandarim)": "zh-CN", "🇧🇷 Português": "pt", "🇪🇸 Español": "es"}
        
        col_lang1, col_lang2 = st.columns(2)
        with col_lang1: idioma_origem = st.selectbox("🗣️ Idioma dos Dados Originais:", ["🇧🇷 Português", "🇺🇸 English"])
        with col_lang2: idioma_destino = st.selectbox("🎯 Traduzir e Formatar para:", ["🇺🇸 English", "🇨🇳 中文 (Mandarim)", "🇪🇸 Español"])
        
        sigla_origem = mapa_idiomas[idioma_origem]
        sigla_destino = mapa_idiomas[idioma_destino]
        
        st.markdown("---")
        
        # A INTELIGÊNCIA DE LINHAS: O usuário diz onde a tabela começa
        linha_inicio = st.number_input("📌 Em qual linha da planilha começam os nomes das colunas? (Ex: Se os dados começam na 13 e os nomes na 12, digite 12)", min_value=1, value=1)
        
        arquivo_dados = st.file_uploader("📁 Arraste a sua Planilha Bruta aqui (Apenas XLSX)", type=["xlsx"])
        
        if arquivo_dados:
            try:
                # O pandas agora ignora o "lixo" do topo (header=0 significa primeira linha, então diminuímos 1)
                df = pd.read_excel(arquivo_dados, header=linha_inicio - 1)
                colunas = list(df.columns)
                
                st.success("Tabela localizada com sucesso!")
                st.markdown("### 🤖 Mapeamento de Colunas")
                st.info("💡 Dica: Na descrição, você pode selecionar múltiplas colunas (ex: Cor, Tamanho, Voltagem). O robô vai juntar tudo em um texto só.")
                
                c1, c2 = st.columns(2)
                with c1: 
                    # MULTISELECT: Permite escolher várias colunas para fundir
                    cols_desc = st.multiselect("✅ Quais colunas formam a Descrição do Produto?", colunas)
                with c2: 
                    # Dropdown normal para a quantidade, com a opção "Não tem"
                    opcoes_qtd = ["Não tem na planilha"] + colunas
                    col_qtd = st.selectbox("✅ Qual é a coluna de Quantidade?", opcoes_qtd)
                
                st.write("<br>", unsafe_allow_html=True)
                st.markdown("<div class='btn-main'>", unsafe_allow_html=True)
                
                if st.button("🚀 TRADUZIR E GERAR PLANILHA CORPORATIVA"):
                    if len(cols_desc) == 0:
                        st.error("Por favor, selecione pelo menos uma coluna para a Descrição do Produto.")
                    else:
                        with st.spinner("O motor SpeedQuote está construindo sua cotação e fundindo as colunas. Aguarde..."):
                            tradutor = GoogleTranslator(source=sigla_origem, target=sigla_destino)
                            
                            wb = openpyxl.Workbook()
                            ws = wb.active
                            ws.title = "SpeedQuote Format"
                            
                            headers = ["Item No.", "Image", "Product Description", "Quantity", "Target Price", "Quoted Price", "MOQ", "Notes"]
                            ws.append(headers)
                            
                            header_fill = PatternFill(start_color="1C86EE", end_color="1C86EE", fill_type="solid")
                            header_font = Font(color="FFFFFF", bold=True)
                            for cell in ws[1]:
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                            
                            ws.column_dimensions['A'].width = 10
                            ws.column_dimensions['B'].width = 15
                            ws.column_dimensions['C'].width = 60
                            ws.column_dimensions['D'].width = 15
                            ws.column_dimensions['E'].width = 15
                            ws.column_dimensions['F'].width = 15
                            ws.column_dimensions['G'].width = 15
                            ws.column_dimensions['H'].width = 30
                            
                            barra_progresso = st.progress(0)
                            total_itens = len(df)
                            
                            for index, row in df.iterrows():
                                # FUSÃO DE COLUNAS: Junta todos os pedaços da descrição
                                partes_desc = []
                                for col in cols_desc:
                                    valor = str(row[col]) if pd.notna(row[col]) else ""
                                    # Limpa valores nulos do pandas
                                    if valor and valor.strip().lower() != "nan":
                                        partes_desc.append(valor)
                                
                                # Junta as partes com um hífen ou espaço
                                texto_original = " - ".join(partes_desc)
                                
                                # Puxa a quantidade
                                quantidade = ""
                                if col_qtd != "Não tem na planilha":
                                    quantidade = str(row[col_qtd]) if pd.notna(row[col_qtd]) else ""
                                    if quantidade.strip().lower() == "nan": quantidade = ""
                                
                                try:
                                    texto_traduzido = tradutor.translate(texto_original[:4500]) if texto_original else ""
                                except:
                                    texto_traduzido = texto_original
                                    
                                ws.append([index + 1, "", texto_traduzido, quantidade, "", "", "", ""])
                                ws.cell(row=index+2, column=3).alignment = Alignment(wrap_text=True, vertical="center")
                                ws.row_dimensions[index+2].height = 40
                                
                                barra_progresso.progress((index + 1) / total_itens)
                            
                            output = io.BytesIO()
                            wb.save(output)
                            output.seek(0)
                            
                            st.success("🎉 Arquivo Corporativo gerado com sucesso!")
                            st.download_button(
                                label="📥 BAIXAR PLANILHA PADRONIZADA",
                                data=output,
                                file_name=f"Cotação_Formatada_{sigla_destino}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
            except Exception as e:
                st.error(f"Ocorreu um erro ao ler a planilha. Verifique a linha indicada. Detalhe técnico: {e}")
        st.markdown("</div>", unsafe_allow_html=True)
